import os
import shutil
from typing import List

import openpyxl
from django.conf import settings

from phasma_food_v2.measurements.models import Measurement

HEADER_NIR = [
    'wave',
    'preprocessed',
    'darkReference',
    'whiteReference'
]
HEADER_FLUO = [
    'wave',
    'preprocessed',
    "rawData1", "rawData2", "rawData3", "rawData4", "rawData5",
    "rawData6", "rawData7", "rawData8", "rawData9", "rawData10",
    "avgData",
    "rawWhite1", "rawWhite2", "rawWhite3", "rawWhite4", "rawWhite5",
    "rawWhite6", "rawWhite7", "rawWhite8", "rawWhite9", "rawWhite10",
    "correctedWhite",
    "avgWhite",
    "rawDark1", "rawDark2", "rawDark3", "rawDark4", "rawDark5",
    "rawDark6", "rawDark7", "rawDark8", "rawDark9", "rawDark10",
    "avgDark",
]
HEADER_VIS = HEADER_FLUO + [
    "rawDarkforWhite1", "rawDarkforWhite2", "rawDarkforWhite3", "rawDarkforWhite4", "rawDarkforWhite5",
    "rawDarkforWhite6", "rawDarkforWhite7", "rawDarkforWhite8", "rawDarkforWhite9", "rawDarkforWhite10"
]

HEADER_FOOD_ADULTERATION = [
    'Food Type', 'Food Subtype', 'Adulteration Sample ID', 'Other species', 'Purity SMP', 'Alcohol label', 'Authentic',
    'Low value filler', 'Nitrogen enhancer', 'Diluted %', 'Hazard 1 name', 'Hazard 1 %', 'Hazard 2 name', 'Hazard 2 %'
]
HEADER_FOOD_SPOILAGE = [
    'Food Type', 'Temperature', 'Temperature Exposure Hours', 'Microbiological SampleID', 'Microbiological Unit',
    'Microbiological Value'
]
HEADER_MYCOTOXINS_DETECTION = [
    'Food Type', 'Mycotoxins', 'Granularity', 'Aflatoxin Name', 'Aflatoxin Unit', 'Aflatoxin Value'
]


def excel_zip(email: str, measurements: List[str]) -> List[str]:
    """Generate excel files and compress them into zip."""
    measurements_all = Measurement.objects.all()
    measurements_filtered = measurements_all.filter(sample_id__in=measurements)
    location = settings.MEDIA_ROOT
    folder_name = email.split("@")[0]
    full_folder_path = "{}/excel/".format(location)
    if not os.path.exists(full_folder_path):
        os.makedirs(full_folder_path, 0o777)
    if not os.path.exists("{}{}/".format(full_folder_path, folder_name)):
        os.makedirs("{}{}/".format(full_folder_path, folder_name), 0o777)

    for measurement in measurements_filtered:
        measurement_data_time = measurement.date_created
        measurement_sample_id = measurement.sample_id
        white_reference = measurements_all.filter(
            use_case="White Reference"
        ).filter(
            date_created__lt=measurement_data_time
        ).order_by(
            '-date_created'
        ).first()

        wb = openpyxl.Workbook()
        ws_sample_info = wb.create_sheet("Sample Info")
        ws_vis = wb.create_sheet("VIS")
        ws_vis.append(HEADER_VIS)
        ws_nir = wb.create_sheet("NIR")
        ws_nir.append(HEADER_NIR)
        ws_fluo = wb.create_sheet("FLUO")
        ws_fluo.append(HEADER_FLUO)
        wb.remove(wb["Sheet"])

        # VIS
        preprocessed_vis = measurement.vis["preprocessed"] if "preprocessed" in measurement.vis else None
        corrected_white_vis = measurement.vis["whiteReference"] if "whiteReference" in measurement.vis else None
        raw_data_vis = measurement.vis["rawData"] if "rawData" in measurement.vis else None
        avg_data_vis = measurement.vis["avgData"] if "avgData" in measurement.vis else None
        raw_white_vis = measurement.vis["rawWhite"] if "rawWhite" in measurement.vis else None
        avg_white_vis = measurement.vis["avgWhite"] if "avgWhite" in measurement.vis else None
        raw_dark_vis = measurement.vis["rawDark"] if "rawDark" in measurement.vis else None
        avg_dark_vis = measurement.vis["avgDark"] if "avgDark" in measurement.vis else None
        raw_dark_vis_wr = white_reference.vis["rawDark"] if white_reference else None
        if raw_data_vis:
            raw_data_01, raw_data_02, raw_data_03, raw_data_04, raw_data_05, \
            raw_data_06, raw_data_07, raw_data_08, raw_data_09, raw_data_10 = raw_data_vis
        else:
            raw_data_01 = raw_data_02 = raw_data_03 = raw_data_04 = raw_data_05 = \
                raw_data_06 = raw_data_07 = raw_data_08 = raw_data_09 = raw_data_10 = None
        if raw_white_vis:
            raw_white_01, raw_white_02, raw_white_03, raw_white_04, raw_white_05, \
            raw_white_06, raw_white_07, raw_white_08, raw_white_09, raw_white_10 = raw_white_vis
        else:
            raw_white_01 = raw_white_02 = raw_white_03 = raw_white_04 = raw_white_05 = \
                raw_white_06 = raw_white_07 = raw_white_08 = raw_white_09 = raw_white_10 = None
        if raw_dark_vis:
            raw_dark_01, raw_dark_02, raw_dark_03, raw_dark_04, raw_dark_05, \
            raw_dark_06, raw_dark_07, raw_dark_08, raw_dark_09, raw_dark_10 = raw_dark_vis
        else:
            raw_dark_01 = raw_dark_02 = raw_dark_03 = raw_dark_04 = raw_dark_05 = \
                raw_dark_06 = raw_dark_07 = raw_dark_08 = raw_dark_09 = raw_dark_10 = None
        if raw_dark_vis_wr:
            raw_dark_wr_01, raw_dark_wr_02, raw_dark_wr_03, raw_dark_wr_04, raw_dark_wr_05, \
            raw_dark_wr_06, raw_dark_wr_07, raw_dark_wr_08, raw_dark_wr_09, raw_dark_wr_10 = raw_dark_vis_wr
        else:
            raw_dark_wr_01 = raw_dark_wr_02 = raw_dark_wr_03 = raw_dark_wr_04 = raw_dark_wr_05 = \
                raw_dark_wr_06 = raw_dark_wr_07 = raw_dark_wr_08 = raw_dark_wr_09 = raw_dark_wr_10 = None
        while preprocessed_vis:
            preprocessed = preprocessed_vis.pop()
            preprocessed_wave, preprocessed_measurement = preprocessed["wave"], preprocessed["measurement"]
            corrected_white = corrected_white_vis.pop()["measurement"] if corrected_white_vis else None
            raw_data_one = raw_data_01.pop()["measurement"] if raw_data_01 else None
            raw_data_two = raw_data_02.pop()["measurement"] if raw_data_02 else None
            raw_data_three = raw_data_03.pop()["measurement"] if raw_data_03 else None
            raw_data_four = raw_data_04.pop()["measurement"] if raw_data_04 else None
            raw_data_five = raw_data_05.pop()["measurement"] if raw_data_05 else None
            raw_data_six = raw_data_06.pop()["measurement"] if raw_data_06 else None
            raw_data_seven = raw_data_07.pop()["measurement"] if raw_data_07 else None
            raw_data_eight = raw_data_08.pop()["measurement"] if raw_data_08 else None
            raw_data_nine = raw_data_09.pop()["measurement"] if raw_data_09 else None
            raw_data_ten = raw_data_10.pop()["measurement"] if raw_data_10 else None
            avg_data = avg_data_vis.pop()["measurement"] if avg_data_vis else None
            raw_white_one = raw_white_01.pop()["measurement"] if raw_white_01 else None
            raw_white_two = raw_white_02.pop()["measurement"] if raw_white_02 else None
            raw_white_three = raw_white_03.pop()["measurement"] if raw_white_03 else None
            raw_white_four = raw_white_04.pop()["measurement"] if raw_white_04 else None
            raw_white_five = raw_white_05.pop()["measurement"] if raw_white_05 else None
            raw_white_six = raw_white_06.pop()["measurement"] if raw_white_06 else None
            raw_white_seven = raw_white_07.pop()["measurement"] if raw_white_07 else None
            raw_white_eight = raw_white_08.pop()["measurement"] if raw_white_08 else None
            raw_white_nine = raw_white_09.pop()["measurement"] if raw_white_09 else None
            raw_white_ten = raw_white_10.pop()["measurement"] if raw_white_10 else None
            avg_white = avg_white_vis.pop()["measurement"] if avg_white_vis else None
            raw_dark_one = raw_dark_01.pop()["measurement"] if raw_dark_01 else None
            raw_dark_two = raw_dark_02.pop()["measurement"] if raw_dark_02 else None
            raw_dark_three = raw_dark_03.pop()["measurement"] if raw_dark_03 else None
            raw_dark_four = raw_dark_04.pop()["measurement"] if raw_dark_04 else None
            raw_dark_five = raw_dark_05.pop()["measurement"] if raw_dark_05 else None
            raw_dark_six = raw_dark_06.pop()["measurement"] if raw_dark_06 else None
            raw_dark_seven = raw_dark_07.pop()["measurement"] if raw_dark_07 else None
            raw_dark_eight = raw_dark_08.pop()["measurement"] if raw_dark_08 else None
            raw_dark_nine = raw_dark_09.pop()["measurement"] if raw_dark_09 else None
            raw_dark_ten = raw_dark_10.pop()["measurement"] if raw_dark_10 else None
            avg_dark = avg_dark_vis.pop()["measurement"] if avg_dark_vis else None
            raw_dark_wr_one = raw_dark_wr_01.pop()["measurement"] if raw_dark_wr_01 else None
            raw_dark_wr_two = raw_dark_wr_02.pop()["measurement"] if raw_dark_wr_02 else None
            raw_dark_wr_three = raw_dark_wr_03.pop()["measurement"] if raw_dark_wr_03 else None
            raw_dark_wr_four = raw_dark_wr_04.pop()["measurement"] if raw_dark_wr_04 else None
            raw_dark_wr_five = raw_dark_wr_05.pop()["measurement"] if raw_dark_wr_05 else None
            raw_dark_wr_six = raw_dark_wr_06.pop()["measurement"] if raw_dark_wr_06 else None
            raw_dark_wr_seven = raw_dark_wr_07.pop()["measurement"] if raw_dark_wr_07 else None
            raw_dark_wr_eight = raw_dark_wr_08.pop()["measurement"] if raw_dark_wr_08 else None
            raw_dark_wr_nine = raw_dark_wr_09.pop()["measurement"] if raw_dark_wr_09 else None
            raw_dark_wr_ten = raw_dark_wr_10.pop()["measurement"] if raw_dark_wr_10 else None
            data_vis = [
                preprocessed_wave, preprocessed_measurement, raw_data_one, raw_data_two, raw_data_three,
                raw_data_four, raw_data_five, raw_data_six, raw_data_seven, raw_data_eight, raw_data_nine,
                raw_data_ten, avg_data, raw_white_one, raw_white_two, raw_white_three, raw_white_four,
                raw_white_five, raw_white_six, raw_white_seven, raw_white_eight, raw_white_nine, raw_white_ten,
                corrected_white, avg_white, raw_dark_one, raw_dark_two, raw_dark_three, raw_dark_four,
                raw_dark_five, raw_dark_six, raw_dark_seven, raw_dark_eight, raw_dark_nine, raw_dark_ten, avg_dark,
                raw_dark_wr_one, raw_dark_wr_two, raw_dark_wr_three, raw_dark_wr_four, raw_dark_wr_five,
                raw_dark_wr_six, raw_dark_wr_seven, raw_dark_wr_eight, raw_dark_wr_nine, raw_dark_wr_ten
            ]
            ws_vis.append(data_vis)

        # NIR
        preprocessed_nir = measurement.nir["preprocessed"] if "preprocessed" in measurement.nir else None
        dark_reference_nir = measurement.nir["darkReference"] if "darkReference" in measurement.nir else None
        white_reference_nir = measurement.nir["whiteReference"] if "whiteReference" in measurement.nir else None
        while preprocessed_nir:
            preprocessed = preprocessed_nir.pop()
            preprocessed_wave, preprocessed_measurement = preprocessed["wave"], preprocessed["measurement"]
            dark_reference = dark_reference_nir.pop()["measurement"] if dark_reference_nir else None
            white_reference = white_reference_nir.pop()["measurement"] if white_reference_nir else None
            data = [preprocessed_wave, preprocessed_measurement, dark_reference, white_reference]
            ws_nir.append(data)

        # FLUO
        preprocessed_fluo = measurement.fluo["preprocessed"] if "preprocessed" in measurement.fluo else None
        raw_data_fluo = measurement.fluo["rawData"] if "rawData" in measurement.fluo else None
        avg_data_fluo = measurement.fluo["avgData"] if "avgData" in measurement.fluo else None
        raw_white_fluo = measurement.fluo["rawWhite"] if "rawWhite" in measurement.fluo else None
        avg_white_fluo = measurement.fluo["avgWhite"] if "avgWhite" in measurement.fluo else None
        raw_dark_fluo = measurement.fluo["rawDark"] if "rawDark" in measurement.fluo else None
        avg_dark_fluo = measurement.fluo["avgDark"] if "avgDark" in measurement.fluo else None
        if len(raw_data_fluo):
            raw_data_fluo_01, raw_data_fluo_02, raw_data_fluo_03, raw_data_fluo_04, raw_data_fluo_05, \
            raw_data_fluo_06, raw_data_fluo_07, raw_data_fluo_08, raw_data_fluo_09, raw_data_fluo_10 = raw_data_fluo
        else:
            raw_data_fluo_01 = raw_data_fluo_02 = raw_data_fluo_03 = raw_data_fluo_04 = raw_data_fluo_05 = \
                raw_data_fluo_06 = raw_data_fluo_07 = raw_data_fluo_08 = raw_data_fluo_09 = raw_data_fluo_10 = None
        if len(raw_white_fluo):
            raw_white_fluo_01, raw_white_fluo_02, raw_white_fluo_03, raw_white_fluo_04, raw_white_fluo_05, \
            raw_white_fluo_06, raw_white_fluo_07, raw_white_fluo_08, raw_white_fluo_09, raw_white_fluo_10 = raw_white_fluo
        else:
            raw_white_fluo_01 = raw_white_fluo_02 = raw_white_fluo_03 = raw_white_fluo_04 = raw_white_fluo_05 = \
                raw_white_fluo_06 = raw_white_fluo_07 = raw_white_fluo_08 = raw_white_fluo_09 = raw_white_fluo_10 = None
        if len(raw_dark_fluo):
            raw_dark_fluo_01, raw_dark_fluo_02, raw_dark_fluo_03, raw_dark_fluo_04, raw_dark_fluo_05, \
            raw_dark_fluo_06, raw_dark_fluo_07, raw_dark_fluo_08, raw_dark_fluo_09, raw_dark_fluo_10 = raw_dark_fluo
        else:
            raw_dark_fluo_01 = raw_dark_fluo_02 = raw_dark_fluo_03 = raw_dark_fluo_04 = raw_dark_fluo_05 = \
                raw_dark_fluo_06 = raw_dark_fluo_07 = raw_dark_fluo_08 = raw_dark_fluo_09 = raw_dark_fluo_10 = None
        while preprocessed_fluo:
            preprocessed = preprocessed_fluo.pop()
            preprocessed_wave, preprocessed_measurement = preprocessed["wave"], preprocessed["measurement"]
            raw_data_one = raw_data_fluo_01.pop()["measurement"] if raw_data_fluo_01 else None
            raw_data_two = raw_data_fluo_02.pop()["measurement"] if raw_data_fluo_02 else None
            raw_data_three = raw_data_fluo_03.pop()["measurement"] if raw_data_fluo_03 else None
            raw_data_four = raw_data_fluo_04.pop()["measurement"] if raw_data_fluo_04 else None
            raw_data_five = raw_data_fluo_05.pop()["measurement"] if raw_data_fluo_05 else None
            raw_data_six = raw_data_fluo_06.pop()["measurement"] if raw_data_fluo_06 else None
            raw_data_seven = raw_data_fluo_07.pop()["measurement"] if raw_data_fluo_07 else None
            raw_data_eight = raw_data_fluo_08.pop()["measurement"] if raw_data_fluo_08 else None
            raw_data_nine = raw_data_fluo_09.pop()["measurement"] if raw_data_fluo_09 else None
            raw_data_ten = raw_data_fluo_10.pop()["measurement"] if raw_data_fluo_10 else None
            avg_data = avg_data_fluo.pop()["measurement"] if avg_data_fluo else None
            raw_white_one = raw_white_fluo_01.pop()["measurement"] if raw_white_fluo_01 else None
            raw_white_two = raw_white_fluo_02.pop()["measurement"] if raw_white_fluo_02 else None
            raw_white_three = raw_white_fluo_03.pop()["measurement"] if raw_white_fluo_03 else None
            raw_white_four = raw_white_fluo_04.pop()["measurement"] if raw_white_fluo_04 else None
            raw_white_five = raw_white_fluo_05.pop()["measurement"] if raw_white_fluo_05 else None
            raw_white_six = raw_white_fluo_06.pop()["measurement"] if raw_white_fluo_06 else None
            raw_white_seven = raw_white_fluo_07.pop()["measurement"] if raw_white_fluo_07 else None
            raw_white_eight = raw_white_fluo_08.pop()["measurement"] if raw_white_fluo_08 else None
            raw_white_nine = raw_white_fluo_09.pop()["measurement"] if raw_white_fluo_09 else None
            raw_white_ten = raw_white_fluo_10.pop()["measurement"] if raw_white_fluo_10 else None
            avg_white = avg_white_fluo.pop()["measurement"] if avg_white_fluo else None
            raw_dark_one = raw_dark_fluo_01.pop()["measurement"] if raw_dark_fluo_01 else None
            raw_dark_two = raw_dark_fluo_02.pop()["measurement"] if raw_dark_fluo_02 else None
            raw_dark_three = raw_dark_fluo_03.pop()["measurement"] if raw_dark_fluo_03 else None
            raw_dark_four = raw_dark_fluo_04.pop()["measurement"] if raw_dark_fluo_04 else None
            raw_dark_five = raw_dark_fluo_05.pop()["measurement"] if raw_dark_fluo_05 else None
            raw_dark_six = raw_dark_fluo_06.pop()["measurement"] if raw_dark_fluo_06 else None
            raw_dark_seven = raw_dark_fluo_07.pop()["measurement"] if raw_dark_fluo_07 else None
            raw_dark_eight = raw_dark_fluo_08.pop()["measurement"] if raw_dark_fluo_08 else None
            raw_dark_nine = raw_dark_fluo_09.pop()["measurement"] if raw_dark_fluo_09 else None
            raw_dark_ten = raw_dark_fluo_10.pop()["measurement"] if raw_dark_fluo_10 else None
            avg_dark = avg_dark_fluo.pop()["measurement"] if avg_dark_fluo else None
            data_fluo = [
                preprocessed_wave, preprocessed_measurement, raw_data_one, raw_data_two, raw_data_three,
                raw_data_four, raw_data_five, raw_data_six, raw_data_seven, raw_data_eight, raw_data_nine,
                raw_data_ten, avg_data, raw_white_one, raw_white_two, raw_white_three, raw_white_four,
                raw_white_five, raw_white_six, raw_white_seven, raw_white_eight, raw_white_nine, raw_white_ten,
                avg_white, raw_dark_one, raw_dark_two, raw_dark_three, raw_dark_four,
                raw_dark_five, raw_dark_six, raw_dark_seven, raw_dark_eight, raw_dark_nine, raw_dark_ten, avg_dark
            ]
            ws_fluo.append(data_fluo)

        if measurement.use_case == 'Food adulteration':
            ws_sample_info.append(HEADER_FOOD_ADULTERATION)
            ws_sample_info.append([
                measurement.food_type, measurement.food_subtype, measurement.adulteration_id, measurement.other_species,
                measurement.purity_smp, measurement.alcohol_label, measurement.authentic, measurement.low_value_filler,
                measurement.nitrogen_enhancer, measurement.diluted_pct, measurement.hazard_one_name,
                measurement.hazard_one_pct, measurement.hazard_two_name, measurement.hazard_two_pct
            ])
            name_part = measurement.adulteration_id
        elif measurement.use_case == "Food spoilage":
            ws_sample_info.append(HEADER_FOOD_SPOILAGE)
            ws_sample_info.append([
                measurement.food_type, measurement.temperature, measurement.temperature_exposure_hours,
                measurement.microbiological_id, measurement.microbiological_unit, measurement.microbiological_value
            ])
            name_part = measurement.microbiological_id
        elif measurement.use_case == 'Mycotoxins detection':
            ws_sample_info.append(HEADER_MYCOTOXINS_DETECTION)
            ws_sample_info.append([
                measurement.food_type, measurement.mycotoxins, measurement.granularity, measurement.aflatoxin_name,
                measurement.aflatoxin_unit, measurement.aflatoxin_value
            ])
            name_part = measurement.food_type
        else:
            name_part = "other"

        time_stamp = str(measurement.date_created).replace(" ", '')

        excel_name = "{}{}/{}-{}-{}.xlsx".format(
            full_folder_path,
            folder_name,
            time_stamp,
            name_part,
            measurement_sample_id
        )
        wb.save("{}".format(excel_name))
        wb.close()

    shutil.make_archive(
        base_name="{}/{}".format(
            full_folder_path,
            folder_name.replace(".", "")
        ),
        format="zip",
        root_dir="{}{}".format(
            full_folder_path, folder_name
        ),
    )
    return [full_folder_path, folder_name]
